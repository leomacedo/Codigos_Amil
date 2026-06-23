import os
import glob
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def processar_base_excel():
    # 1. Localizar o arquivo contendo "base" no nome na mesma pasta do script
    arquivos = glob.glob("*base*.xlsx") + glob.glob("*base*.xls")
    
    if not arquivos:
        print("Erro: Nenhum arquivo contendo a palavra 'base' foi encontrado na pasta.")
        return
    
    caminho_arquivo = arquivos[0]
    print(f"Arquivo encontrado: {caminho_arquivo}")
    
    # 2. Carregar os dados da planilha original
    df = pd.read_excel(caminho_arquivo)
    
    # Mapeamento para busca de colunas independente de espaços extras ou diferenças de caixa
    colunas_originais = {col.strip().lower(): col for col in df.columns}
    
    jornada_col = colunas_originais.get('jornada')
    origem_col = colunas_originais.get('origem')
    data_col = colunas_originais.get('data de entrada na linha')
    
    if not all([jornada_col, origem_col, data_col]):
        print("Erro: Verifique se as colunas 'Jornada', 'Origem' e 'Data de entra na linha' existem no arquivo.")
        return
        
    # 3. Limpeza e preparação dos dados essenciais
    df['data_datetime'] = pd.to_datetime(df[data_col], errors='coerce')
    df = df.dropna(subset=['data_datetime'])
    
    # Ordenar por data para garantir as abas da mais antiga para a mais recente
    df = df.sort_values(by='data_datetime')
    
    # Tratar texto das colunas para garantir o casamento perfeito com a lista fixa
    # Remove espaços extras nas pontas e garante a comparação limpa
    df[jornada_col] = df[jornada_col].astype(str).str.strip()
    df['origem_clean'] = df[origem_col].astype(str).str.strip()
    
    # Criar chaves de agrupamento de tempo
    df['ano_mes_key'] = df['data_datetime'].dt.to_period('M')
    df['dia'] = df['data_datetime'].dt.day
    df['nome_aba'] = df['data_datetime'].dt.strftime('%m-%Y')
    
    # --- ORDEM FIXA DAS JORNADAS SOLICITADA ---
    ordem_jornadas_fixa = [
        "Anticoagulante Seguro",
        "Cuidado Cardíaco Valvar",
        "Cuidado Integral da Mama",
        "Cuidado Oncológico Colorretal",
        "Cuidado Oncológico Próstata",
        "Cuidado Oncológico Pulmonar",
        "Cuidados para Endometriose",
        "Cuidados Pós Infarto",
        "Emagrecimento",
        "Fumo Zero",
        "Gestação Segura",
        "Insuficiência Cardíaca Controlada",
        "Pós AVC",
        "Ritmo Certo",
        "Saúde da Coluna",
        "Saúde Mental",
        "Saúde Renal"
    ]
    
    # Criar um mapeamento temporário em lowercase para evitar problemas caso a base original tenha divergências de maiúsculas/minúsculas
    jornadas_lower = [j.lower() for j in ordem_jornadas_fixa]
    mapa_jornadas_originais = dict(zip(jornadas_lower, ordem_jornadas_fixa))
    
    # Ajusta a coluna da base de dados para bater exatamente com a grafia da sua lista fixa
    df['jornada_ajustada'] = df[jornada_col].str.lower().map(mapa_jornadas_originais).fillna(df[jornada_col])
    
    # Obter os períodos únicos na ordem cronológica correta
    periodos_ordenados = df['ano_mes_key'].unique()
    arquivo_saida = "Consolidado_Mensal.xlsx"
    
    # --- CONFIGURAÇÃO DE DESIGN E ESTILOS ---
    fonte_titulo_secao = Font(name='Calibri', size=11, bold=True, color='1F4E78')
    fonte_cabecalho_tabela = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    fonte_corpo = Font(name='Calibri', size=11, bold=False)
    fonte_total = Font(name='Calibri', size=11, bold=True)
    
    fill_titulo_secao = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    fill_cabecalho = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    fill_zebra = PatternFill(start_color='F9FBFD', end_color='F9FBFD', fill_type='solid')
    fill_total = PatternFill(start_color='E9EEF4', end_color='E9EEF4', fill_type='solid')
    
    alinhar_esquerda = Alignment(horizontal='left', vertical='center')
    alinhar_centro = Alignment(horizontal='center', vertical='center')
    
    borda_fina = Side(border_style="thin", color="D9D9D9")
    borda_dupla = Side(border_style="double", color="808080")
    layout_borda_celula = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)
    layout_borda_total = Border(top=borda_fina, bottom=borda_dupla)
    
    # 4. Construindo a nova planilha usando o ExcelWriter
    with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
        
        for periodo in periodos_ordenados:
            group_mes = df[df['ano_mes_key'] == periodo]
            nome_aba = group_mes['nome_aba'].iloc[0]
            
            workbook = writer.book
            ws = workbook.create_sheet(title=nome_aba)
            ws.views.sheetView[0].showGridLines = True
            
            linha_atual = 2
            
            # Agrupar e listar as origens dentro do mês atual
            origens_no_mes = sorted(group_mes['origem_clean'].unique())
            
            # Descobrir todos os dias existentes para este mês específico (evita misturar colunas vazias globais)
            todos_dias_mes = sorted(list(group_mes['dia'].unique()))
            
            for origem in origens_no_mes:
                group_origem = group_mes[group_mes['origem_clean'] == origem]
                
                # Criar tabela dinâmica (Pivot Table)
                tabela_dinamica = group_origem.pivot_table(
                    index='jornada_ajustada',
                    columns='dia',
                    aggfunc='size',
                    fill_value=0
                )
                
                # Reindexar as linhas usando a ORDEM FIXA e as colunas usando TODOS OS DIAS do mês
                tabela_dinamica = tabela_dinamica.reindex(index=ordem_jornadas_fixa, columns=todos_dias_mes, fill_value=0)
                
                # Título da Origem (Ex: AMIL 360)
                ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=len(todos_dias_mes) + 2)
                celula_titulo = ws.cell(row=linha_atual, column=1, value=origem.upper())
                celula_titulo.font = fonte_titulo_secao
                celula_titulo.fill = fill_titulo_secao
                celula_titulo.alignment = alinhar_esquerda
                linha_atual += 1
                
                # Cabeçalhos da Tabela (Jornada, Dias, SOMA)
                headers = ['Jornada'] + [str(d) for d in todos_dias_mes] + ['SOMA']
                for col_idx, text in enumerate(headers, start=1):
                    cell = ws.cell(row=linha_atual, column=col_idx, value=text)
                    cell.font = fonte_cabecalho_tabela
                    cell.fill = fill_cabecalho
                    cell.alignment = alinhar_centro if col_idx > 1 else alinhar_esquerda
                linha_atual += 1
                
                # Preenchimento das Linhas (Garantindo a lista estática e cor alternada)
                contador_linhas = 0
                for jornada_nome, r_data in tabela_dinamica.iterrows():
                    cell_jornada = ws.cell(row=linha_atual, column=1, value=jornada_nome)
                    cell_jornada.font = fonte_corpo
                    cell_jornada.alignment = alinhar_esquerda
                    cell_jornada.border = layout_borda_celula
                    
                    soma_linha = 0
                    for idx_d, d in enumerate(todos_dias_mes, start=2):
                        v_dia = r_data[d]
                        soma_linha += v_dia
                        cell_v = ws.cell(row=linha_atual, column=idx_d, value=v_dia)
                        cell_v.font = fonte_corpo
                        cell_v.alignment = alinhar_centro
                        cell_v.border = layout_borda_celula
                        if contador_linhas % 2 == 1:
                            cell_v.fill = fill_zebra
                            cell_jornada.fill = fill_zebra
                            
                    # Coluna de SOMA
                    cell_soma = ws.cell(row=linha_atual, column=len(todos_dias_mes) + 2, value=soma_linha)
                    cell_soma.font = fonte_total
                    cell_soma.alignment = alinhar_centro
                    cell_soma.border = layout_borda_celula
                    if contador_linhas % 2 == 1:
                        cell_soma.fill = fill_zebra
                        
                    linha_atual += 1
                    contador_linhas += 1
                
                # Linha de Totais Gerais por Coluna
                cell_tot_label = ws.cell(row=linha_atual, column=1, value="Total Geral")
                cell_tot_label.font = fonte_total
                cell_tot_label.fill = fill_total
                cell_tot_label.border = layout_borda_total
                
                for idx_d, d in enumerate(todos_dias_mes, start=2):
                    col_sum = tabela_dinamica[d].sum()
                    cell_col_sum = ws.cell(row=linha_atual, column=idx_d, value=col_sum)
                    cell_col_sum.font = fonte_total
                    cell_col_sum.fill = fill_total
                    cell_col_sum.alignment = alinhar_centro
                    cell_col_sum.border = layout_borda_total
                    
                total_geral_origem = tabela_dinamica.values.sum()
                cell_total_final = ws.cell(row=linha_atual, column=len(todos_dias_mes) + 2, value=total_geral_origem)
                cell_total_final.font = fonte_total
                cell_total_final.fill = fill_total
                cell_total_final.alignment = alinhar_centro
                cell_total_final.border = layout_borda_total
                
                linha_atual += 4
            
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
                
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 10)
                
    print(f"Processamento concluído com sucesso! Planilha gerada: {arquivo_saida}")

if __name__ == '__main__':
    processar_base_excel()
