import os
import sys
import re
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


ARQUIVO_ORIGEM = "baseorigem.xlsx"
ARQUIVO_SAIDA = "base_formatada.xlsx"
ABA_ORIGEM = None

COLUNA_JORNADA = "Jornada"

REGRAS_ORIGEM = {
    "Amil 360": "Captado pela Amil/Call Center",
    "Forms Eric (Desosp Total Care)": "Forms"
}

POSICOES_JORNADA = {
    "anticoagulante seguro": (38, 39),
    "emagrecimento": (42, 43)
}


# Define a pasta base do script ou executável
if getattr(sys, "frozen", False):
    diretorio_atual = os.path.dirname(sys.executable)
else:
    diretorio_atual = os.path.dirname(os.path.abspath(__file__))

arquivo_origem = os.path.join(diretorio_atual, ARQUIVO_ORIGEM)
arquivo_saida = os.path.join(diretorio_atual, ARQUIVO_SAIDA)


# Busca uma coluna considerando nomes possíveis
def buscar_coluna(colunas_normalizadas, *nomes_possiveis):
    for nome in nomes_possiveis:
        chave = str(nome).strip().lower()
        if chave in colunas_normalizadas:
            return colunas_normalizadas[chave]
    return None


# Puxa o valor da linha pela primeira coluna encontrada
def puxar_valor(linha, colunas_normalizadas, *nomes_possiveis):
    coluna = buscar_coluna(colunas_normalizadas, *nomes_possiveis)

    if coluna is None:
        return ""

    valor = linha[coluna]

    if pd.isna(valor):
        return ""

    return valor


# Formata datas no padrão brasileiro
def formatar_data_brasileira(valor):
    if pd.isna(valor) or valor == "":
        return ""

    data = pd.to_datetime(valor, errors="coerce", dayfirst=True)

    if pd.isna(data):
        return ""

    return data.strftime("%d/%m/%Y")


# Limpa o nome da aba para o Excel
def limpar_nome_aba(nome):
    nome = str(nome).strip()

    if nome == "" or nome.lower() == "nan":
        nome = "Sem Jornada"

    nome = re.sub(r"[\[\]\:\*\?\/\\]", "-", nome)
    nome = nome[:31]

    return nome


# Trata a origem conforme regras definidas
def tratar_origem(valor):
    if pd.isna(valor) or valor == "":
        return ""

    valor = str(valor).strip()

    return REGRAS_ORIGEM.get(valor, valor)


# Retorna as posições das colunas conforme a jornada
def posicoes_por_jornada(jornada):
    jornada_tratada = str(jornada).strip().lower()

    if jornada_tratada in POSICOES_JORNADA:
        return POSICOES_JORNADA[jornada_tratada]

    return 36, 37


# Carrega a planilha de origem
df_origem = pd.read_excel(arquivo_origem, sheet_name=ABA_ORIGEM, engine="openpyxl")

if isinstance(df_origem, dict):
    primeira_aba = list(df_origem.keys())[0]
    df_origem = df_origem[primeira_aba]

df_origem.columns = df_origem.columns.astype(str).str.strip()
colunas_normalizadas = {str(col).strip().lower(): col for col in df_origem.columns}

col_jornada = buscar_coluna(colunas_normalizadas, COLUNA_JORNADA)

if col_jornada is None:
    raise Exception(f"A coluna '{COLUNA_JORNADA}' não foi encontrada na planilha de origem.")

data_hoje = datetime.today().strftime("%d/%m/%Y")

df_origem["_Jornada_Tratada"] = df_origem[col_jornada].fillna("Sem Jornada").astype(str).str.strip()
df_origem.loc[df_origem["_Jornada_Tratada"] == "", "_Jornada_Tratada"] = "Sem Jornada"

wb = Workbook()
ws_inicial = wb.active
wb.remove(ws_inicial)

abas_criadas = set()

# Cria uma aba para cada jornada
for jornada in df_origem["_Jornada_Tratada"].unique():
    df_filtrado = df_origem[df_origem["_Jornada_Tratada"] == jornada].copy()

    nome_aba = limpar_nome_aba(jornada)
    nome_aba_original = nome_aba
    contador = 1

    while nome_aba in abas_criadas:
        sufixo = f"_{contador}"
        nome_aba = nome_aba_original[:31 - len(sufixo)] + sufixo
        contador += 1

    abas_criadas.add(nome_aba)

    ws = wb.create_sheet(title=nome_aba)

    col_flag, col_encaminhado = posicoes_por_jornada(jornada)
    col_final_inicio = col_encaminhado + 1

    cabecalhos_fixos = {
        1: "Prioridade no Contato?",
        2: "Data Envio Paciente",
        3: "Motivo Envio",
        4: "Origem",
        5: "MO",
        6: "Nome do paciente",
        7: "CPF",
        8: "Data de Nascimento",
        9: "Idade",
        10: "Genero",
        11: "Estado",
        12: "Município",
        13: "Bairro",
        14: "Nível de Rede",
        15: "Telefone atualizado",
        col_flag: "Flag",
        col_encaminhado: "Encaminhado p/ Consolidado?"
    }

    cabecalhos_finais = [
        "Linha",
        "Mês/Ano",
        "Semana",
        "Mês/Ano Inclusão",
        "Semana Inclusão",
        "SLA Captação",
        "SLA Exclusão",
        "SLA Tent. Contatos",
        "SLA Consulta",
        "Ativo/Inativo"
    ]

    # Preenche os cabeçalhos fixos
    for coluna, cabecalho in cabecalhos_fixos.items():
        ws.cell(row=1, column=coluna, value=cabecalho)

    # Preenche os cabeçalhos finais
    for i, cabecalho in enumerate(cabecalhos_finais):
        ws.cell(row=1, column=col_final_inicio + i, value=cabecalho)

    linha_excel = 2

    # Preenche as linhas da aba
    for _, linha in df_filtrado.iterrows():
        ws.cell(row=linha_excel, column=1, value="Sim")
        ws.cell(row=linha_excel, column=2, value=data_hoje)
        ws.cell(row=linha_excel, column=3, value="Navegação")
        ws.cell(row=linha_excel, column=4, value=tratar_origem(puxar_valor(linha, colunas_normalizadas, "Origem")))
        ws.cell(row=linha_excel, column=5, value=puxar_valor(linha, colunas_normalizadas, "MO"))
        ws.cell(row=linha_excel, column=6, value=puxar_valor(linha, colunas_normalizadas, "Nome do paciente", "Nome Paciente", "Paciente"))
        ws.cell(row=linha_excel, column=7, value=puxar_valor(linha, colunas_normalizadas, "CPF"))
        ws.cell(row=linha_excel, column=8, value=formatar_data_brasileira(puxar_valor(linha, colunas_normalizadas, "Data de Nascimento", "Nascimento")))
        ws.cell(row=linha_excel, column=9, value="")
        ws.cell(row=linha_excel, column=10, value=puxar_valor(linha, colunas_normalizadas, "Genero", "Gênero"))
        ws.cell(row=linha_excel, column=11, value=puxar_valor(linha, colunas_normalizadas, "Estado", "UF"))
        ws.cell(row=linha_excel, column=12, value=puxar_valor(linha, colunas_normalizadas, "Município", "Municipio", "Cidade"))
        ws.cell(row=linha_excel, column=13, value=puxar_valor(linha, colunas_normalizadas, "Bairro"))
        ws.cell(row=linha_excel, column=14, value=puxar_valor(linha, colunas_normalizadas, "Nível de Rede", "Nivel de Rede"))
        ws.cell(row=linha_excel, column=15, value=puxar_valor(linha, colunas_normalizadas, "Telefone atualizado", "Telefone", "Celular"))
        ws.cell(row=linha_excel, column=col_flag, value="Sim")
        ws.cell(row=linha_excel, column=col_encaminhado, value="Sim")

        linha_excel += 1

    # Ajusta largura das colunas
    for coluna in range(1, ws.max_column + 1):
        letra = get_column_letter(coluna)
        ws.column_dimensions[letra].width = 22

    ws.freeze_panes = "A2"

wb.save(arquivo_saida)

print(f"Arquivo gerado com sucesso: {arquivo_saida}")