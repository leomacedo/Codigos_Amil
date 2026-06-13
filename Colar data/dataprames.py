import pandas as pd
from openpyxl import load_workbook

# =========================
# CONFIGURAÇÕES
# =========================

arquivo_origem = "datas.xlsx"
arquivo_saida = "datas_com_mes.xlsx"

# Nome da coluna onde estão as datas
coluna_datas = "Data de Entrada na Linha"

# =========================
# LER PLANILHA
# =========================

df = pd.read_excel(arquivo_origem, engine="openpyxl")

# =========================
# MESES EM PORTUGUÊS
# =========================

meses = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Março",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro"
}

# =========================
# FUNÇÃO PARA TRANSFORMAR DATA EM TEXTO MÊS/ANO
# =========================

def identificar_mes_ano(valor):
    if pd.isna(valor) or str(valor).strip() == "":
        return ""

    data = pd.to_datetime(valor, errors="coerce", dayfirst=True)

    if pd.isna(data):
        return ""

    mes = meses[data.month]
    ano = data.year

    return f"{mes} {ano}"

# =========================
# CRIAR COLUNA AO LADO COMO TEXTO
# =========================

df["Mês/Ano"] = df[coluna_datas].apply(identificar_mes_ano).astype(str)

# =========================
# SALVAR ARQUIVO
# =========================

df.to_excel(arquivo_saida, index=False, engine="openpyxl")

# =========================
# FORÇAR A COLUNA MÊS/ANO COMO TEXTO NO EXCEL
# =========================

wb = load_workbook(arquivo_saida)
ws = wb.active

for col in ws[1]:
    if col.value == "Mês/Ano":
        letra_coluna = col.column_letter

        for linha in range(2, ws.max_row + 1):
            celula = ws[f"{letra_coluna}{linha}"]

            if celula.value is not None and str(celula.value).strip() != "":
                celula.value = str(celula.value)

            celula.number_format = "@"

        break

wb.save(arquivo_saida)

print(f"Arquivo gerado com sucesso: {arquivo_saida}")